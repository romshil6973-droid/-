"""
Тесты модуля reporter.py — генерация Excel-отчётов.
"""

import pytest
import sys
import os
import tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from datetime import datetime
from pathlib import Path

import openpyxl

from core.database import Database
from core.reporter import ExcelReporter, _fmt_dur, _fmt_time, _date_ru


# ─── Утилиты форматирования ───────────────────────────────────────────────────

class TestFormatters:
    def test_fmt_dur_zero(self):
        assert _fmt_dur(0) == "00:00:00"

    def test_fmt_dur_none(self):
        assert _fmt_dur(None) == "00:00:00"

    def test_fmt_dur_hour(self):
        assert _fmt_dur(3661) == "01:01:01"

    def test_fmt_time_valid(self):
        assert _fmt_time("2024-06-10 09:30:00") == "09:30:00"

    def test_fmt_time_none(self):
        assert _fmt_time(None) == "—"

    def test_fmt_time_invalid(self):
        result = _fmt_time("invalid")
        assert result == "invalid"

    def test_date_ru(self):
        dt = datetime(2024, 6, 10)  # Понедельник
        result = _date_ru(dt)
        assert "10" in result
        assert "июня" in result
        assert "2024" in result
        assert "понедельник" in result


# ─── Генерация Excel ──────────────────────────────────────────────────────────

@pytest.fixture
def reporter_env():
    """Создаёт DB + Reporter + временную папку для отчётов."""
    db = Database(db_path=":memory:")
    reporter = ExcelReporter(db)
    with tempfile.TemporaryDirectory() as tmpdir:
        # Путь для Google Drive: {tmpdir}/{login}/Отчёт_{date}.xlsx
        yield db, reporter, tmpdir


LOGIN = "testuser"


class TestExcelGeneration:
    def test_creates_file(self, reporter_env):
        db, reporter, tmpdir = reporter_env
        path = reporter.generate("2024-06-10", LOGIN, tmpdir)
        assert Path(path).exists()

    def test_filename_contains_date_only(self, reporter_env):
        """Имя файла должно содержать только дату — без ФИО."""
        db, reporter, tmpdir = reporter_env
        path = reporter.generate("2024-06-10", LOGIN, tmpdir)
        name = Path(path).name
        assert "2024-06-10" in name
        # ФИО не должно быть в имени файла
        assert "Иванов" not in name
        assert "shilov" not in name

    def test_file_saved_in_login_subfolder(self, reporter_env):
        """Файл должен быть в {gdrive_path}/{login}/"""
        db, reporter, tmpdir = reporter_env
        path = reporter.generate("2024-06-10", LOGIN, tmpdir)
        assert Path(path).parent.name == LOGIN

    def test_two_sheets(self, reporter_env):
        db, reporter, tmpdir = reporter_env
        path = reporter.generate("2024-06-10", LOGIN, tmpdir)
        wb = openpyxl.load_workbook(path)
        assert "Фото рабочего дня" in wb.sheetnames
        assert "Мониторинг активности" in wb.sheetnames

    def test_login_in_sheet_no_fio(self, reporter_env):
        """В отчёте должен быть логин, а ФИО и должность — отсутствовать."""
        db, reporter, tmpdir = reporter_env
        path = reporter.generate("2024-06-10", LOGIN, tmpdir)
        wb = openpyxl.load_workbook(path)
        ws = wb["Фото рабочего дня"]
        all_text = " ".join(
            str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        )
        assert LOGIN in all_text, "Логин не найден в отчёте"
        assert "Должность" not in all_text, "Поле 'Должность' не должно быть в отчёте"
        assert "Сотрудник:" not in all_text, "Поле 'Сотрудник' не должно быть в отчёте"

    def test_work_sessions_in_sheet(self, reporter_env):
        db, reporter, tmpdir = reporter_env
        sid = db.start_work_session("Акты сверок", datetime(2024, 6, 10, 9, 0))
        db.complete_work_session(sid, datetime(2024, 6, 10, 9, 45))
        path = reporter.generate("2024-06-10", LOGIN, tmpdir)

        wb = openpyxl.load_workbook(path)
        ws = wb["Фото рабочего дня"]
        found = any(
            "Акты сверок" in str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        )
        assert found, "Операция не найдена в листе Excel"

    def test_activity_in_sheet(self, reporter_env):
        db, reporter, tmpdir = reporter_env
        aid = db.log_activity_start(
            "chrome", "Google Chrome",
            datetime(2024, 6, 10, 9, 0)
        )
        db.log_activity_end(aid, datetime(2024, 6, 10, 9, 30))
        path = reporter.generate("2024-06-10", LOGIN, tmpdir)

        wb = openpyxl.load_workbook(path)
        ws = wb["Мониторинг активности"]
        found = any(
            "chrome" in str(cell.value or "").lower()
            for row in ws.iter_rows()
            for cell in row
        )
        assert found, "Активность не найдена в листе Excel"

    def test_empty_data_no_crash(self, reporter_env):
        """Генерация отчёта без данных не должна падать."""
        db, reporter, tmpdir = reporter_env
        path = reporter.generate("2024-01-01", LOGIN, tmpdir)
        assert Path(path).exists()

    def test_achievement_in_report(self, reporter_env):
        db, reporter, tmpdir = reporter_env
        db.save_achievement("Выполнил план на 120%", "2024-06-10")
        path = reporter.generate("2024-06-10", LOGIN, tmpdir)

        wb = openpyxl.load_workbook(path)
        ws = wb["Фото рабочего дня"]
        found = any(
            "120%" in str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        )
        assert found, "Достижение не найдено в отчёте"


# ─── Тесты модуля config.py ───────────────────────────────────────────────────

class TestConfig:
    def test_validate_login_ok(self):
        from core.config import validate_login
        ok, _ = validate_login("shilov")
        assert ok

    def test_validate_login_with_numbers(self):
        from core.config import validate_login
        ok, _ = validate_login("user123")
        assert ok

    def test_validate_login_empty(self):
        from core.config import validate_login
        ok, msg = validate_login("")
        assert not ok
        assert msg

    def test_validate_login_cyrillic(self):
        from core.config import validate_login
        ok, msg = validate_login("шилов")
        assert not ok

    def test_validate_login_too_short(self):
        from core.config import validate_login
        ok, _ = validate_login("ab")
        assert not ok

    def test_validate_login_too_long(self):
        from core.config import validate_login
        ok, _ = validate_login("a" * 21)
        assert not ok

    def test_validate_login_uppercase(self):
        from core.config import validate_login
        ok, _ = validate_login("Shilov")
        assert not ok

    def test_validate_gdrive_path_nonexistent(self):
        from core.config import validate_gdrive_path
        ok, msg = validate_gdrive_path(r"C:\nonexistent\path\xyz123")
        assert not ok

    def test_validate_gdrive_path_empty(self):
        from core.config import validate_gdrive_path
        ok, msg = validate_gdrive_path("")
        assert not ok

    def test_validate_gdrive_path_valid(self, tmp_path):
        from core.config import validate_gdrive_path
        ok, msg = validate_gdrive_path(str(tmp_path))
        assert ok

    def test_save_and_load_config(self, tmp_path, monkeypatch):
        """Сохранение и загрузка конфига."""
        from core import config as cfg_module
        config_file = tmp_path / "config.ini"
        monkeypatch.setattr(cfg_module, "get_config_path", lambda: config_file)

        cfg_module.save_config("shilov", r"C:\GDrive\Reports")
        result = cfg_module.load_config()

        assert result["login"] == "shilov"
        assert result["google_drive_path"] == r"C:\GDrive\Reports"

    def test_config_exists_false_when_no_file(self, tmp_path, monkeypatch):
        from core import config as cfg_module
        config_file = tmp_path / "no_config.ini"
        monkeypatch.setattr(cfg_module, "get_config_path", lambda: config_file)
        assert not cfg_module.config_exists()

    def test_config_exists_true_after_save(self, tmp_path, monkeypatch):
        from core import config as cfg_module
        config_file = tmp_path / "config.ini"
        monkeypatch.setattr(cfg_module, "get_config_path", lambda: config_file)
        cfg_module.save_config("shilov", r"C:\GDrive\Reports")
        assert cfg_module.config_exists()
