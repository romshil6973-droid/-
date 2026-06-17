"""
Tests for browser_history.py module.

Covers:
  - is_browser() - browser process detection
  - get_urls_in_window() - URL filtering by time window
  - load_browser_history() - correct behavior when no browsers installed
  - Chrome timestamp conversion
  - Firefox timestamp conversion
  - Deduplication and filtering of internal URLs
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import sqlite3
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

import pytest

from core.browser_history import (
    is_browser,
    get_urls_in_window,
    load_browser_history,
    _chrome_ts_to_datetime,
    _firefox_ts_to_datetime,
    _read_chromium_history,
)


# --- is_browser --------------------------------------------------------------

class TestIsBrowser:
    def test_chrome(self):
        assert is_browser("chrome") is True

    def test_msedge(self):
        assert is_browser("msedge") is True

    def test_firefox(self):
        assert is_browser("firefox") is True

    def test_yandex_browser(self):
        # Yandex Browser appears as "browser" in process list
        assert is_browser("browser") is True

    def test_word_not_browser(self):
        assert is_browser("WINWORD") is False

    def test_excel_not_browser(self):
        assert is_browser("EXCEL") is False

    def test_python_not_browser(self):
        assert is_browser("python") is False

    def test_case_insensitive(self):
        assert is_browser("Chrome") is True
        assert is_browser("CHROME") is True


# --- Timestamp conversion ----------------------------------------------------

class TestTimestampConversion:
    def test_chrome_ts_returns_datetime(self):
        # ~2012 in Chrome microseconds since 1601-01-01
        ts = 13_000_000_000_000_000
        result = _chrome_ts_to_datetime(ts)
        assert isinstance(result, datetime)
        assert result.year > 2000

    def test_firefox_ts_returns_datetime(self):
        # ~2023 in Firefox microseconds since Unix epoch
        ts = 1_700_000_000_000_000
        result = _firefox_ts_to_datetime(ts)
        assert isinstance(result, datetime)
        assert result.year >= 2023

    def test_chrome_ts_zero_is_epoch(self):
        # 0 = 1601-01-01, function must not raise
        result = _chrome_ts_to_datetime(0)
        assert isinstance(result, datetime)


# --- get_urls_in_window ------------------------------------------------------

class TestGetUrlsInWindow:
    def _make_history(self):
        base = datetime(2024, 6, 10, 10, 0, 0)
        return [
            (base + timedelta(minutes=0),  "https://google.com"),
            (base + timedelta(minutes=5),  "https://github.com"),
            (base + timedelta(minutes=10), "https://stackoverflow.com"),
            (base + timedelta(minutes=20), "https://youtube.com"),
        ]

    def test_returns_urls_in_window(self):
        history = self._make_history()
        start = datetime(2024, 6, 10, 10, 0, 0)
        end   = datetime(2024, 6, 10, 10, 12, 0)
        result = get_urls_in_window(history, start, end)
        assert "https://google.com" in result
        assert "https://github.com" in result
        assert "https://stackoverflow.com" in result
        assert "https://youtube.com" not in result  # outside window

    def test_empty_window(self):
        history = self._make_history()
        start = datetime(2024, 6, 10, 9, 0, 0)
        end   = datetime(2024, 6, 10, 9, 59, 0)
        result = get_urls_in_window(history, start, end)
        assert result == []

    def test_empty_history(self):
        result = get_urls_in_window([], datetime.now(), datetime.now())
        assert result == []

    def test_exact_boundary_included(self):
        history = [(datetime(2024, 6, 10, 10, 0, 0), "https://exact.com")]
        result = get_urls_in_window(
            history,
            datetime(2024, 6, 10, 10, 0, 0),
            datetime(2024, 6, 10, 10, 0, 0)
        )
        assert "https://exact.com" in result

    def test_no_duplicates(self):
        ts = datetime(2024, 6, 10, 10, 0, 0)
        history = [
            (ts,                            "https://same.com"),
            (ts + timedelta(seconds=30),    "https://same.com"),
            (ts + timedelta(seconds=60),    "https://other.com"),
        ]
        result = get_urls_in_window(history, ts, ts + timedelta(minutes=2))
        assert result.count("https://same.com") == 1
        assert "https://other.com" in result

    def test_order_preserved(self):
        history = self._make_history()
        start = datetime(2024, 6, 10, 10, 0, 0)
        end   = datetime(2024, 6, 10, 10, 15, 0)
        result = get_urls_in_window(history, start, end)
        assert result[0] == "https://google.com"
        assert result[-1] == "https://stackoverflow.com"


# --- load_browser_history ----------------------------------------------------

class TestLoadBrowserHistory:
    def test_returns_list_when_no_browser_installed(self):
        """Returns empty list when no browser profiles exist; does not raise."""
        result = load_browser_history("1970-01-01")
        assert isinstance(result, list)

    def test_filters_internal_urls(self):
        """Internal URLs (chrome://, about:, data:) must be excluded."""
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "History"
            conn = sqlite3.connect(str(db_path))
            conn.execute("""
                CREATE TABLE urls (
                    id INTEGER PRIMARY KEY,
                    url TEXT,
                    title TEXT,
                    last_visit_time INTEGER
                )
            """)
            conn.execute("""
                CREATE TABLE visits (
                    id INTEGER PRIMARY KEY,
                    url INTEGER,
                    visit_time INTEGER
                )
            """)
            from core.browser_history import _CHROME_EPOCH
            from datetime import timezone
            import datetime as dt_module

            def to_chrome_ts(d):
                epoch = _CHROME_EPOCH
                delta = d.replace(tzinfo=timezone.utc) - epoch
                return int(delta.total_seconds() * 1_000_000)

            test_date = datetime(2024, 6, 10, 10, 0, 0)

            urls_data = [
                (1, "chrome://settings", "Settings",    to_chrome_ts(test_date)),
                (2, "about:blank",        "Blank",       to_chrome_ts(test_date)),
                (3, "https://google.com", "Google",      to_chrome_ts(test_date)),
                (4, "data:text/html,...", "Data",        to_chrome_ts(test_date)),
                (5, "edge://flags",       "Edge Flags",  to_chrome_ts(test_date)),
            ]
            visits_data = [(i, i, to_chrome_ts(test_date + timedelta(seconds=i*10)))
                           for i in range(1, 6)]

            conn.executemany("INSERT INTO urls VALUES (?,?,?,?)", urls_data)
            conn.executemany("INSERT INTO visits VALUES (?,?,?)", visits_data)
            conn.commit()
            conn.close()

            result = _read_chromium_history(db_path, "2024-06-10")
            urls_only = [url for _, url in result]

            assert "https://google.com" in urls_only
            assert "chrome://settings" not in urls_only
            assert "about:blank" not in urls_only
            assert "data:text/html,..." not in urls_only
            assert "edge://flags" not in urls_only

    def test_read_chromium_history_real_data(self):
        """Reads Chrome-format SQLite and returns correct URL + datetime."""
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "History"
            conn = sqlite3.connect(str(db_path))
            conn.execute("""
                CREATE TABLE urls (
                    id INTEGER PRIMARY KEY, url TEXT,
                    title TEXT, last_visit_time INTEGER
                )
            """)
            conn.execute("""
                CREATE TABLE visits (
                    id INTEGER PRIMARY KEY,
                    url INTEGER, visit_time INTEGER
                )
            """)

            from core.browser_history import _CHROME_EPOCH
            from datetime import timezone

            def to_chrome_ts(d):
                delta = d.replace(tzinfo=timezone.utc) - _CHROME_EPOCH
                return int(delta.total_seconds() * 1_000_000)

            test_dt = datetime(2024, 6, 10, 9, 30, 0)
            ts = to_chrome_ts(test_dt)

            conn.execute("INSERT INTO urls VALUES (1, 'https://example.com', 'Example', ?)", (ts,))
            conn.execute("INSERT INTO visits VALUES (1, 1, ?)", (ts,))
            conn.commit()
            conn.close()

            result = _read_chromium_history(db_path, "2024-06-10")
            assert len(result) == 1
            dt_result, url = result[0]
            assert url == "https://example.com"
            assert isinstance(dt_result, datetime)
            assert dt_result.strftime('%Y-%m-%d') == "2024-06-10"

    def test_read_chromium_history_missing_file(self):
        """Non-existent history file returns empty list."""
        result = _read_chromium_history(Path("/nonexistent/History"), "2024-06-10")
        assert result == []

    def test_read_chromium_history_wrong_date(self):
        """URLs from a different date are not returned."""
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "History"
            conn = sqlite3.connect(str(db_path))
            conn.execute("CREATE TABLE urls (id INTEGER PRIMARY KEY, url TEXT, title TEXT, last_visit_time INTEGER)")
            conn.execute("CREATE TABLE visits (id INTEGER PRIMARY KEY, url INTEGER, visit_time INTEGER)")

            from core.browser_history import _CHROME_EPOCH
            from datetime import timezone

            test_dt = datetime(2024, 6, 10, 9, 30, 0)
            delta = test_dt.replace(tzinfo=timezone.utc) - _CHROME_EPOCH
            ts = int(delta.total_seconds() * 1_000_000)

            conn.execute("INSERT INTO urls VALUES (1, 'https://example.com', 'Example', ?)", (ts,))
            conn.execute("INSERT INTO visits VALUES (1, 1, ?)", (ts,))
            conn.commit()
            conn.close()

            result = _read_chromium_history(db_path, "2024-06-11")
            assert result == []


# --- Integration: reporter + browser_history ---------------------------------

class TestReporterWithUrlColumn:
    """Verify the report generates correctly with the new URL column."""

    def test_activity_sheet_has_url_column(self):
        """Activity sheet must contain the header 'URL (browser)'."""
        import openpyxl, tempfile
        from core.database import Database
        from core.reporter import ExcelReporter

        db = Database(db_path=":memory:")
        reporter = ExcelReporter(db)

        with tempfile.TemporaryDirectory() as tmpdir:
            path = reporter.generate("2024-06-10", "testuser", tmpdir)
            wb = openpyxl.load_workbook(path)

            # Sheet name contains Cyrillic - access by index
            ws = wb.worksheets[1]  # second sheet = activity

            # Find the column-headers row (first column = numero sign U+2116)
            header_row = None
            for row in ws.iter_rows(values_only=True):
                if row[0] == "№":
                    header_row = list(row)
                    break

            assert header_row is not None, "Headers row not found in sheet"
            assert "URL (браузер)" in header_row, \
                f"URL column not found. Headers: {header_row}"

    def test_activity_sheet_has_six_columns(self):
        """Activity sheet must have 6 columns (was 5)."""
        import openpyxl, tempfile
        from core.database import Database
        from core.reporter import ExcelReporter

        db = Database(db_path=":memory:")
        aid = db.log_activity_start("chrome", "Google Chrome", datetime(2024, 6, 10, 9, 0))
        db.log_activity_end(aid, datetime(2024, 6, 10, 9, 30))

        reporter = ExcelReporter(db)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = reporter.generate("2024-06-10", "testuser", tmpdir)
            wb = openpyxl.load_workbook(path)
            ws = wb.worksheets[1]

            header_row = None
            for row in ws.iter_rows(values_only=True):
                if row[0] == "№":
                    header_row = row
                    break

            assert header_row is not None, "Headers row not found"
            assert len([c for c in header_row if c is not None]) == 6

    def test_browser_row_url_empty_when_no_history(self):
        """URL cell is empty (no crash) when no browser history exists."""
        import openpyxl, tempfile
        from core.database import Database
        from core.reporter import ExcelReporter

        db = Database(db_path=":memory:")
        aid = db.log_activity_start("browser", "Yandex", datetime(2024, 6, 10, 9, 0))
        db.log_activity_end(aid, datetime(2024, 6, 10, 9, 30))

        reporter = ExcelReporter(db)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = reporter.generate("2024-06-10", "testuser", tmpdir)
            assert Path(path).exists()
