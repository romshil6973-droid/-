"""
Генератор отчётов Excel.
Лист 1: Фото рабочего дня
Лист 2: Мониторинг активности
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .database import Database
from .browser_history import load_browser_history, get_urls_in_window, is_browser

# ─── Цвета (брендбук "Система Порядка") ───────────────────────────────────────
C_TITLE_BG   = "0B1C3D"   # Тёмно-синий — строка заголовка отчёта
C_HEADER_BG  = "1E3A5F"   # Синий — заголовки колонок таблицы
C_ROW_EVEN   = "EBF5FF"   # Светло-голубой — чётные строки
C_ROW_ODD    = "FFFFFF"   # Белый — нечётные строки
C_TOTAL_BG   = "D6E8FF"   # Голубой — строка итогов
C_WHITE      = "FFFFFF"
C_ACCENT     = "3B82F6"   # Синий акцент (текст)


def _border() -> Border:
    """Тонкая рамка ячейки."""
    s = Side(style="thin", color="B0C4DE")
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt_time(ts: str | None) -> str:
    """Извлекает ЧЧ:ММ:СС из строки 'YYYY-MM-DD HH:MM:SS'."""
    if not ts:
        return "—"
    try:
        return datetime.strptime(ts, '%Y-%m-%d %H:%M:%S').strftime('%H:%M:%S')
    except ValueError:
        return ts


def _fmt_dur(seconds: int | None) -> str:
    """Форматирует секунды → ЧЧ:ММ:СС."""
    if not seconds:
        return "00:00:00"
    h, rem = divmod(int(seconds), 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def _date_ru(dt: datetime) -> str:
    """10 июня 2026, среда."""
    months = ["января","февраля","марта","апреля","мая","июня",
              "июля","августа","сентября","октября","ноября","декабря"]
    days   = ["понедельник","вторник","среда","четверг","пятница","суббота","воскресенье"]
    return f"{dt.day} {months[dt.month-1]} {dt.year}, {days[dt.weekday()]}"


class ExcelReporter:
    """Создаёт Excel-отчёт за указанную дату."""

    def __init__(self, db: Database):
        self.db = db

    def generate(self, date_str: str, report_dir: str, employee: dict) -> str:
        """
        Генерирует Excel-файл.

        Args:
            date_str:   Дата в формате YYYY-MM-DD
            report_dir: Папка для сохранения
            employee:   {'name': ..., 'surname': ..., 'position': ...}

        Returns:
            Полный путь к созданному файлу.
        """
        date_obj  = datetime.strptime(date_str, '%Y-%m-%d')
        work_rows = self.db.get_work_sessions_for_date(date_str)
        act_rows  = self.db.get_activity_for_date(date_str)
        day_info  = self.db.get_day_session(date_str)
        achieve   = self.db.get_achievement(date_str)

        wb = openpyxl.Workbook()

        self._sheet_workday(wb, date_obj, work_rows, day_info, achieve, employee)
        self._sheet_activity(wb, date_obj, act_rows, employee)

        # Путь файла
        Path(report_dir).mkdir(parents=True, exist_ok=True)
        fio = f"{employee.get('surname','')} {employee.get('name','')}".strip() or "Сотрудник"
        filename = f"Отчёт_{fio}_{date_str}.xlsx"
        filepath = str(Path(report_dir) / filename)
        wb.save(filepath)
        return filepath

    # ──────────────────────────────────────────────────────────────────────────
    # Приватные методы
    # ──────────────────────────────────────────────────────────────────────────

    def _title_row(self, ws, text: str, cols: int = 5):
        """Записывает строку-заголовок отчёта (тёмно-синяя полоса)."""
        ws.append([text])
        row = ws.max_row
        ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
        cell = ws.cell(row, 1)
        cell.fill = PatternFill("solid", fgColor=C_TITLE_BG)
        cell.font = Font(bold=True, color=C_WHITE, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 38

    def _meta_row(self, ws, date_ru: str, fio: str, position: str, cols: int = 5):
        """Строка с датой, ФИО, должностью под заголовком."""
        ws.append([f"Дата: {date_ru}", "", f"Сотрудник: {fio}", "", f"Должность: {position}"])
        row = ws.max_row
        fill = PatternFill("solid", fgColor="14294D")
        font = Font(color=C_WHITE, size=11)
        for c in range(1, cols + 1):
            cell = ws.cell(row, c)
            cell.fill = fill
            cell.font = font
        ws.row_dimensions[row].height = 22

    def _header_row(self, ws, headers: list) -> int:
        """Записывает заголовки таблицы (синий фон). Возвращает номер строки."""
        ws.append(headers)
        row = ws.max_row
        fill = PatternFill("solid", fgColor=C_HEADER_BG)
        font = Font(bold=True, color=C_WHITE, size=11)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row, c)
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = _border()
        ws.row_dimensions[row].height = 28
        return row

    def _data_row(self, ws, values: list, idx: int, center_cols: set = None):
        """Записывает строку данных с зеброй."""
        ws.append(values)
        row = ws.max_row
        bg = C_ROW_EVEN if idx % 2 == 0 else C_ROW_ODD
        fill = PatternFill("solid", fgColor=bg)
        center_cols = center_cols or {1, 2, 3, 4}
        for c, _ in enumerate(values, 1):
            cell = ws.cell(row, c)
            cell.fill = fill
            cell.font = Font(size=10)
            cell.border = _border()
            cell.alignment = Alignment(
                horizontal="center" if c in center_cols else "left",
                vertical="center",
                wrap_text=c not in center_cols
            )

    def _total_row(self, ws, label: str, value: str, cols: int = 5):
        """Строка итогов."""
        row_data = [""] * cols
        row_data[cols - 3] = label
        row_data[cols - 2] = value
        ws.append(row_data)
        row = ws.max_row
        fill = PatternFill("solid", fgColor=C_TOTAL_BG)
        for c in range(1, cols + 1):
            cell = ws.cell(row, c)
            cell.fill = fill
            cell.font = Font(bold=True, size=11)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ──────────────────────────────────────────────────────────────────────────

    def _sheet_workday(self, wb, date_obj, sessions, day_info, achieve, emp):
        """Лист «Фото рабочего дня»."""
        ws = wb.active
        ws.title = "Фото рабочего дня"

        fio      = f"{emp.get('surname','')} {emp.get('name','')}".strip()
        position = emp.get('position', '')
        date_ru  = _date_ru(date_obj)

        self._title_row(ws, "ФОТО РАБОЧЕГО ДНЯ")
        self._meta_row(ws, date_ru, fio, position)

        # Время начала/конца рабочего дня
        if day_info:
            ws.append([])
            ws.append([
                "Начало рабочего дня:", _fmt_time(day_info.get('work_start')),
                "", "Окончание рабочего дня:", _fmt_time(day_info.get('work_end'))
            ])
            for c in range(1, 6):
                ws.cell(ws.max_row, c).font = Font(bold=True, size=11)

        ws.append([])

        # Таблица операций
        self._header_row(ws, ["№", "Начало", "Конец", "Продолжительность", "Операция"])

        total_sec = 0
        for i, s in enumerate(sessions, 1):
            dur = s.get('duration_seconds') or 0
            total_sec += dur
            self._data_row(ws, [
                i,
                _fmt_time(s.get('start_time')),
                _fmt_time(s.get('end_time')),
                _fmt_dur(dur),
                s.get('operation', '')
            ], i)

        self._total_row(ws, "ИТОГО:", _fmt_dur(total_sec))

        # Достижения
        if achieve and achieve.get('content'):
            ws.append([])
            ws.append(["МОИ ДОСТИЖЕНИЯ ЗА ДЕНЬ:"])
            ws.cell(ws.max_row, 1).font = Font(bold=True, size=12, color=C_ACCENT)
            ws.append([achieve['content']])
            ws.cell(ws.max_row, 1).font = Font(size=11)

        # Ширина столбцов
        for col, width in zip("ABCDE", [5, 12, 12, 18, 55]):
            ws.column_dimensions[col].width = width

    def _sheet_activity(self, wb, date_obj, records, emp):
        """Лист «Мониторинг активности»."""
        ws = wb.create_sheet("Мониторинг активности")

        fio      = f"{emp.get('surname','')} {emp.get('name','')}".strip()
        position = emp.get('position', '')
        date_ru  = _date_ru(date_obj)
        date_str = date_obj.strftime('%Y-%m-%d')

        self._title_row(ws, "МОНИТОРИНГ АКТИВНОСТИ КОМПЬЮТЕРА", cols=6)
        self._meta_row(ws, date_ru, fio, position, cols=6)
        ws.append([])

        self._header_row(ws, ["№", "Начало", "Конец", "Продолжительность",
                               "Приложение / Операция", "URL (браузер)"])

        # Загружаем историю браузеров один раз для всего дня
        try:
            browser_history = load_browser_history(date_str)
        except Exception:
            browser_history = []

        total_active = 0
        for i, r in enumerate(records, 1):
            s_dt_str = r.get('start_time')
            e_dt_str = r.get('end_time')

            # Парсим datetime для сравнения с историей браузера
            try:
                s_dt = datetime.strptime(s_dt_str, '%Y-%m-%d %H:%M:%S') if s_dt_str else None
                e_dt = datetime.strptime(e_dt_str, '%Y-%m-%d %H:%M:%S') if e_dt_str else None
            except ValueError:
                s_dt = e_dt = None

            dur = 0
            if s_dt and e_dt:
                dur = max(0, int((e_dt - s_dt).total_seconds()))

            if r.get('activity_type') != 'idle':
                total_active += dur

            # Читабельное описание приложения
            app   = r.get('app_name', '')
            title = r.get('window_title', '')
            desc  = app
            if title and title != app and len(title) < 90:
                desc = f"{app} — {title}"

            # URL из истории браузера (только для браузерных строк)
            url_text = ""
            if s_dt and e_dt and is_browser(app) and browser_history:
                urls = get_urls_in_window(browser_history, s_dt, e_dt)
                if urls:
                    # Не более 5 URL в одной ячейке, разделённых переносом
                    url_text = "\n".join(urls[:5])
                    if len(urls) > 5:
                        url_text += f"\n... ещё {len(urls) - 5}"

            self._data_row(ws, [
                i, _fmt_time(s_dt_str), _fmt_time(e_dt_str), _fmt_dur(dur), desc, url_text
            ], i, center_cols={1, 2, 3, 4})

        self._total_row(ws, "АКТИВНОЕ ВРЕМЯ:", _fmt_dur(total_active), cols=6)

        # Ширина столбцов
        for col, width in zip("ABCDEF", [5, 12, 12, 18, 55, 60]):
            ws.column_dimensions[col].width = width

        # Высота строк с URL — увеличиваем для читаемости
        for row in ws.iter_rows(min_row=4):
            url_cell = row[5] if len(row) > 5 else None
            if url_cell and url_cell.value and '\n' in str(url_cell.value):
                line_count = str(url_cell.value).count('\n') + 1
                ws.row_dimensions[url_cell.row].height = max(15 * line_count, 30)
